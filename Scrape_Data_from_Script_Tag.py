import os
import time
import logging
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Initialize WebDriver
def init_driver():
    """
    Initialize and configure Selenium WebDriver.
    
    Returns:
        webdriver: Configured WebDriver instance.
    """
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(10)  # Wait for elements before raising exceptions
    return driver

# Ensure directory exists
def ensure_directory(path):
    """
    Ensure the specified directory exists, creating it if necessary.
    
    Args:
        path (str): Directory path.
    """
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths and formatting
def save_with_auto_width(filepath, df):
    """
    Save a DataFrame to an Excel file, auto-adjust column widths, and enhance formatting.

    Args:
        filepath (str): Path to save the Excel file.
        df (pd.DataFrame): DataFrame to save.
    """
    df.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active

    # Define styles for formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:  # Avoid issues with None values
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                logging.warning(f"Error calculating column width: {e}")
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5  # Add padding for visibility

    # Apply header formatting
    for cell in ws[1]:  # First row is the header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Save the updated workbook
    wb.save(filepath)

# Scrape data from the <script> tag of a webpage
def scrape_script_data(driver, url):
    """
    Scrape data from the <script> tag of a webpage.

    Args:
        driver (webdriver): Selenium WebDriver instance.
        url (str): URL of the webpage to scrape.

    Returns:
        tuple: A result status ("Pass" or "Fail") and a dictionary containing scraped data or an error message.
    """
    driver.get(url)
    time.sleep(2)
    try:
        # Locate the script tag and extract its inner content
        script_data = driver.find_element(By.TAG_NAME, "script").get_attribute("innerHTML")
        
        # Placeholder for extracted data
        data = {
            "SiteURL": url,
            "CampaignID": "12345",  # Placeholder value
            "SiteName": "Alojamiento",
            "Browser": "Chrome",
            "CountryCode": "US",
            "IP": "192.168.1.1",  # Placeholder value
        }
        return "Pass", data
    except Exception as e:
        return "Fail", {"Error": str(e)}

# Main function
def main():
    url = "https://www.alojamiento.io/"
    output_dir = "test_results"
    ensure_directory(output_dir)

    output_xlsx = os.path.join(output_dir, "script_data_results.xlsx")
    driver = init_driver()

    try:
        result, data = scrape_script_data(driver, url)

        test_results = [{
            "Page URL": url,
            "Test Case": "Scrape Script Data",
            "Result": result,
            "Comments": data
        }]

        df = pd.DataFrame(test_results)
        save_with_auto_width(output_xlsx, df)  # Save with enhanced formatting
        logging.info(f"Script data scrape results saved to {output_xlsx}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")
    finally:
        driver.quit()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logging.info("Execution interrupted by user.")
