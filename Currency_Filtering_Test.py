import os
import logging
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Initialize WebDriver
def init_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(10)  # Wait for elements before raising exceptions
    return driver

# Ensure directory exists
def ensure_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths
def save_with_auto_width(filepath, df):
    """
    Save a DataFrame to an Excel file, auto-adjust column widths, and enhance formatting.

    Args:
        filepath (str): Path to save the Excel file.
        df (pd.DataFrame): DataFrame to save.
    """
    df.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active

    # Define styles for formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:  # Avoid issues with None values
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                logging.warning(f"Error calculating column width: {e}")
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5  # Add padding for visibility

    # Apply header formatting
    for cell in ws[1]:  # First row is the header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Save the updated workbook
    wb.save(filepath)

# Test currency filter functionality
def test_currency_filter(driver, url):
    logging.info(f"Starting Currency Filter Test for URL: {url}")
    testcase = "Currency Filter Test"
    results = []  # List to store individual test results for each currency

    try:
        # Navigate to the page
        driver.get(url)
        logging.info("Page loaded successfully.")

        # Wait for the page to load completely
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        # Scroll down to load all content
        for _ in range(3):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)

        # Try to locate and click the currency dropdown
        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "js-currency-sort-footer"))
        )
        dropdown.click()
        logging.info("Currency dropdown opened.")

        # Locate the dropdown options
        options = dropdown.find_elements(By.CSS_SELECTOR, ".select-ul > li")
        logging.info(f"Found {len(options)} currency options.")

        # Parse dropdown options
        currency_options = []
        for option in options:
            data_country = option.get_attribute("data-currency-country")
            currency_element = option.find_element(By.CSS_SELECTOR, ".option > p")
            currency_symbol = currency_element.text.split(" ")[0].strip()
            currency_options.append({"country": data_country, "symbol": currency_symbol})
            logging.info(f"Currency option: {data_country} -> {currency_symbol}")

        if not currency_options:
            logging.warning("No currency options found in the dropdown.")
            return [{"Currency Name": "All", "Currency Symbol": "N/A", "Status": "Fail", "Reason": "No currency options found"}]

        # Test each currency
        for currency in currency_options:
            logging.info(f"Testing currency: {currency['country']} -> {currency['symbol']}")

            try:
                # Reopen dropdown
                dropdown = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "js-currency-sort-footer"))
                )
                dropdown.click()

                # Select the currency
                option = next(
                    (opt for opt in options if opt.get_attribute("data-currency-country") == currency["country"]), None
                )
                if not option:
                    logging.warning(f"Option for {currency['country']} not found.")
                    results.append({"Currency Name": currency["country"], "Currency Symbol": currency["symbol"], "Status": "Fail", "Reason": "Currency option not found in dropdown"})
                    continue

                driver.execute_script("arguments[0].scrollIntoView();", option)
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(option)).click()
                time.sleep(2)

                # Validate property tiles
                tiles = driver.find_elements(By.CLASS_NAME, "js-price-value")
                if not tiles:
                    logging.warning(f"No property tiles found for {currency['symbol']}.")
                    results.append({"Currency Name": currency["country"], "Currency Symbol": currency["symbol"], "Status": "Fail", "Reason": "No property tiles found"})
                    continue

                if not all(currency["symbol"] in tile.text for tile in tiles):
                    logging.warning(f"Currency {currency['symbol']} not reflected in property tiles.")
                    results.append({"Currency Name": currency["country"], "Currency Symbol": currency["symbol"], "Status": "Fail", "Reason": f"Currency symbol {currency['symbol']} not found in property tiles"})
                    continue

                logging.info(f"Currency {currency['symbol']} validated successfully.")
                results.append({"Currency Name": currency["country"], "Currency Symbol": currency["symbol"], "Status": "Pass", "Reason": "Validation successful"})

            except Exception as e:
                # Log the full exception details
                logging.error(f"Error testing currency {currency['symbol']}: {str(e)}")
                results.append({"Currency Name": currency["country"], "Currency Symbol": currency["symbol"], "Status": "Fail", "Reason": f"Exception: {str(e)}"})

        return results

    except Exception as e:
        # Log the full exception details for the overall test case
        logging.error(f"Error during {testcase}: {str(e)}")
        return [{"Currency Name": "All", "Currency Symbol": "N/A", "Status": "Fail", "Reason": f"Exception: {str(e)}"}]


# Main function
def main():
    url = "https://www.alojamiento.io/"  # Replace with the actual URL
    output_dir = "test_results"
    output_results_xlsx = os.path.join(output_dir, "currency_test_results.xlsx")
    output_summary_xlsx = os.path.join(output_dir, "currency_test_summary.xlsx")

    # Ensure output directory exists
    ensure_directory(output_dir)

    driver = init_driver()
    try:
        # Run the test
        results = test_currency_filter(driver, url)

        # Save the detailed results to an Excel file with enhanced formatting
        df_results = pd.DataFrame(results)
        save_with_auto_width(output_results_xlsx, df_results)  # Apply the enhanced formatting function

        # Create and save the summary
        fail_results = [res for res in results if res["Status"] == "Fail"]
        pass_count = len([res for res in results if res["Status"] == "Pass"])
        fail_count = len(fail_results)

        summary_data = {
            "Total Currencies Tested": len(results),
            "Pass Count": pass_count,
            "Fail Count": fail_count,
            "Failed Currencies": ", ".join([f"{res['Currency Name']} ({res['Currency Symbol']}) - {res['Reason']}" for res in fail_results])
        }

        # Convert the summary dictionary into a DataFrame
        df_summary = pd.DataFrame([summary_data])

        # Save the summary to an Excel file with enhanced formatting
        save_with_auto_width(output_summary_xlsx, df_summary)  # Apply the enhanced formatting function

        # Log success messages
        logging.info(f"Test results saved to {output_results_xlsx}")
        logging.info(f"Test summary saved to {output_summary_xlsx}")

    except Exception as e:
        # Log any exceptions encountered during execution
        logging.error(f"Error in main execution: {e}")
    finally:
        # Ensure the WebDriver quits regardless of test outcomes
        driver.quit()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logging.info("Execution interrupted by user.")
