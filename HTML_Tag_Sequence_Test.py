import os
import time
import logging
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Initialize WebDriver
def init_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(10)  # Wait for elements before raising exceptions
    return driver

# Ensure directory exists
def ensure_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths and formatting
def save_with_auto_width(filepath, df):
    """
    Save a DataFrame to an Excel file, auto-adjust column widths, and enhance formatting.

    Args:
        filepath (str): Path to save the Excel file.
        df (pd.DataFrame): DataFrame to save.
    """
    df.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active

    # Define styles for formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:  # Avoid issues with None values
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                logging.warning(f"Error calculating column width: {e}")
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5  # Add padding for visibility

    # Apply header formatting
    for cell in ws[1]:  # First row is the header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Save the updated workbook
    wb.save(filepath)

# Test: Check HTML Tag Sequence
def check_html_sequence(driver, url):
    logging.info(f"Starting HTML Tag Sequence Test for URL: {url}")
    driver.get(url)
    time.sleep(2)

    # Find all header tags (h1 to h6)
    headers = driver.find_elements(By.XPATH, "//h1 | //h2 | //h3 | //h4 | //h5 | //h6")
    
    # Collecting header tags with their associated text
    header_info = [{"Tag": header.tag_name.upper(), "Text": header.text} for header in headers]
    levels = [int(header.tag_name[1]) for header in headers]  # Extract numeric levels

    # Log the header information
    for header in header_info:
        logging.info(f"Header Found - Tag: {header['Tag']}, Text: {header['Text']}")

    # Check if the sequence is strictly increasing
    is_valid_sequence = all(x <= y for x, y in zip(levels, levels[1:]))
    
    if is_valid_sequence:
        return "Pass", "HTML tag sequence is valid.", header_info, levels

    # If sequence is broken, return Fail and show the sequence
    return "Fail", f"HTML tag sequence is broken. Found sequence: {levels}", header_info, levels

# Main function
def main():
    url = "https://www.alojamiento.io/"
    
    # Output file paths (updated to new names)
    output_dir = "test_results"
    ensure_directory(output_dir)
    output_xlsx_summary = os.path.join(output_dir, "html_tag_sequence_summary.xlsx")
    output_xlsx_results = os.path.join(output_dir, "html_tag_header_results.xlsx")
    
    driver = init_driver()

    try:
        # Run HTML sequence check and get headers info and sequence
        result, comment, header_info, levels = check_html_sequence(driver, url)

        # Save the test result to the first Excel file
        test_results = [{
            "Page URL": url,
            "Test Case": "HTML Tag Sequence Test",
            "Result": result,
            "Comments": comment,
            "Sequence": str(levels)
        }]
        
        df_results = pd.DataFrame(test_results)
        save_with_auto_width(output_xlsx_summary, df_results)
        logging.info(f"Test results for HTML Tag Sequence Test saved to {output_xlsx_summary}")

        # Save header tag information to the second Excel file
        header_data = [{"Tag": header["Tag"], "Text": header["Text"]} for header in header_info]
        
        # If sequence is broken, also show the correct sequence
        correct_sequence = sorted(levels)
        if result == "Fail":
            header_data.append({"Tag": "Correct Sequence", "Text": str(correct_sequence)})
        
        df_header_info = pd.DataFrame(header_data)
        save_with_auto_width(output_xlsx_results, df_header_info)
        logging.info(f"Header tag information saved to {output_xlsx_results}")

    except Exception as e:
        logging.error(f"Error in main execution: {e}")
    finally:
        driver.quit()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logging.info("Execution interrupted by user.")
