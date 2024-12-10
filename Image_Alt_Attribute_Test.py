import os
import logging
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Initialize WebDriver
def init_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(10)  # Wait for elements before raising exceptions
    return driver

# Ensure directory exists
def ensure_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths and formatting
def save_with_auto_width(filepath, df):
    """
    Save a DataFrame to an Excel file, auto-adjust column widths, and enhance formatting.

    Args:
        filepath (str): Path to save the Excel file.
        df (pd.DataFrame): DataFrame to save.
    """
    df.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active

    # Define styles for formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:  # Avoid issues with None values
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                logging.warning(f"Error calculating column width: {e}")
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5  # Add padding for visibility

    # Apply header formatting
    for cell in ws[1]:  # First row is the header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Save the updated workbook
    wb.save(filepath)

# Test: Check Image Alt Attributes and Save Results
def check_image_alt_and_save(driver, url, output_xlsx, output_summary_xlsx):
    logging.info(f"Starting Image Alt Attribute Test for URL: {url}")
    driver.get(url)
    time.sleep(2)

    # Find all image elements on the page
    images = driver.find_elements(By.TAG_NAME, "img")

    # List to store image attributes and status
    image_data = []
    pass_count = 0
    fail_count = 0

    for index, img in enumerate(images):
        img_src = img.get_attribute("src")  # Get image source
        img_alt = img.get_attribute("alt")  # Get image alt attribute

        # Determine status (Pass or Fail)
        status = "Pass" if img_alt else "Fail"
        
        if status == "Pass":
            pass_count += 1
        else:
            fail_count += 1

        # Append data to the list
        image_data.append({
            "Image Index": index + 1,
            "Image Source": img_src if img_src else "No Source",
            "Alt Text": img_alt if img_alt else "None",
            "Status": status
        })

        # Log the status of each image
        logging.info(f"Image {index + 1}: Source: {img_src}, Alt Text: {img_alt}, Status: {status}")

    # Create a DataFrame for detailed results and save to Excel
    df = pd.DataFrame(image_data)
    save_with_auto_width(output_xlsx, df)
    logging.info(f"Image alt attribute analysis saved to {output_xlsx}")

    # Create a summary DataFrame for Pass/Fail count
    summary_data = [{
        "Test": "Image Alt Text Test Summary",
        "Pass Count": pass_count,
        "Fail Count": fail_count,
        "Total Images": len(images)
    }]
    df_summary = pd.DataFrame(summary_data)
    save_with_auto_width(output_summary_xlsx, df_summary)
    logging.info(f"Image alt attribute summary saved to {output_summary_xlsx}")

# Main function
def main():
    url = "https://www.alojamiento.io/"
    output_dir = "test_results"
    ensure_directory(output_dir)

    output_xlsx = os.path.join(output_dir, "image_alt_results.xlsx")  # Detailed results in .xlsx
    output_summary_xlsx = os.path.join(output_dir, "image_alt_summary.xlsx")  # Summary file in .xlsx
    
    driver = init_driver()

    try:
        check_image_alt_and_save(driver, url, output_xlsx, output_summary_xlsx)
    except Exception as e:
        logging.error(f"An error occurred during execution: {e}")
    finally:
        driver.quit()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logging.info("Execution interrupted by user.")
